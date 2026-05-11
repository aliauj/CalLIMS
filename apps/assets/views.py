from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from .models import Instrument, InstrumentCategory
from apps.clients.models import Client


def _build_export_qs(request):
    """Apply all export filters from request GET params and return (qs, summary_parts)."""
    today = timezone.now().date()
    qs = Instrument.objects.select_related('client', 'category').order_by('asset_tag')

    if request.user.is_client:
        try:
            qs = qs.filter(client=request.user.client_profile)
        except Exception:
            qs = qs.none()

    preset = request.GET.get('preset', '')
    status_filter = request.GET.get('status', '')
    client_filter = request.GET.get('client', '')
    tag_filter = request.GET.get('tag', '').strip()
    desc_filter = request.GET.get('desc', '').strip()
    serial_filter = request.GET.get('serial', '').strip()
    sort_by = request.GET.get('sort', 'tag')

    summary = []

    # Preset overrides
    if preset == 'calibrated':
        qs = qs.filter(status='CALIBRATED')
        summary.append('Status: Calibrated')
    elif preset == 'active':
        qs = qs.filter(status='ACTIVE')
        summary.append('Status: Active / In Service')
    elif preset == 'overdue':
        qs = qs.filter(next_calibration_date__lt=today).exclude(status__in=['OUT_OF_SERVICE', 'DISPOSED'])
        summary.append('Filter: Overdue calibration')
    elif preset == 'due_30':
        qs = qs.filter(
            next_calibration_date__gte=today,
            next_calibration_date__lte=today + timezone.timedelta(days=30),
        ).exclude(status__in=['OUT_OF_SERVICE', 'DISPOSED'])
        summary.append('Filter: Due within 30 days')
    elif preset == 'due_60':
        qs = qs.filter(
            next_calibration_date__gte=today,
            next_calibration_date__lte=today + timezone.timedelta(days=60),
        ).exclude(status__in=['OUT_OF_SERVICE', 'DISPOSED'])
        summary.append('Filter: Due within 60 days')
    elif preset == 'out_of_service':
        qs = qs.filter(status='OUT_OF_SERVICE')
        summary.append('Status: Out of Service')
    elif preset == 'in_calibration':
        qs = qs.filter(status='IN_CALIBRATION')
        summary.append('Status: In Calibration')
    else:
        # Apply individual filters (current view)
        if status_filter:
            qs = qs.filter(status=status_filter)
            summary.append(f'Status: {status_filter}')
        if client_filter == 'internal':
            qs = qs.filter(client__isnull=True)
            summary.append('Customer: Internal')
        elif client_filter:
            qs = qs.filter(client_id=client_filter)
            try:
                summary.append(f'Customer: {Client.objects.get(pk=client_filter).name}')
            except Client.DoesNotExist:
                pass
        if tag_filter:
            qs = qs.filter(asset_tag__icontains=tag_filter)
            summary.append(f'Tag contains: {tag_filter}')
        if desc_filter:
            qs = qs.filter(description__icontains=desc_filter)
            summary.append(f'Description: {desc_filter}')
        if serial_filter:
            qs = qs.filter(serial_number__icontains=serial_filter)
            summary.append(f'Serial: {serial_filter}')

    # Sort
    sort_map = {
        'tag': 'asset_tag',
        'client': 'client__name',
        'due_date': 'next_calibration_date',
        'status': 'status',
        'description': 'description',
    }
    qs = qs.order_by(sort_map.get(sort_by, 'asset_tag'))

    if not summary:
        summary.append('All instruments')
    return qs, summary, today


@login_required
def instrument_list(request):
    qs = Instrument.objects.select_related('client', 'category').order_by('asset_tag')

    if request.user.is_client:
        try:
            qs = qs.filter(client=request.user.client_profile)
        except Exception:
            qs = qs.none()

    # Filters
    status_filter = request.GET.get('status', '')
    client_filter = request.GET.get('client', '')
    tag_filter = request.GET.get('tag', '').strip()
    desc_filter = request.GET.get('desc', '').strip()
    serial_filter = request.GET.get('serial', '').strip()

    if status_filter:
        qs = qs.filter(status=status_filter)
    if client_filter == 'internal':
        qs = qs.filter(client__isnull=True)
    elif client_filter:
        qs = qs.filter(client_id=client_filter)
    if tag_filter:
        qs = qs.filter(asset_tag__icontains=tag_filter)
    if desc_filter:
        qs = qs.filter(description__icontains=desc_filter)
    if serial_filter:
        qs = qs.filter(serial_number__icontains=serial_filter)

    clients = Client.objects.filter(is_active=True).order_by('name') if not request.user.is_client else []

    return render(request, 'instruments/instrument_list.html', {
        'instruments': qs,
        'status_filter': status_filter,
        'client_filter': client_filter,
        'tag_filter': tag_filter,
        'desc_filter': desc_filter,
        'serial_filter': serial_filter,
        'status_choices': Instrument.Status.choices,
        'clients': clients,
    })


@login_required
def instrument_detail(request, pk):
    obj = get_object_or_404(
        Instrument.objects.select_related('client', 'category', 'created_by'), pk=pk
    )
    jobs = obj.calibration_jobs.select_related('method').order_by('-created_at')[:10]
    return render(request, 'instruments/instrument_detail.html', {'instrument': obj, 'jobs': jobs})


@login_required
def instrument_edit(request, pk):
    obj = get_object_or_404(Instrument, pk=pk)
    if request.method == 'POST':
        asset_tag = request.POST.get('asset_tag', '').strip()
        if not asset_tag:
            messages.error(request, 'Instrument Tag cannot be blank.')
            return redirect('instruments:instrument_edit', pk=pk)
        if Instrument.objects.filter(asset_tag=asset_tag).exclude(pk=pk).exists():
            messages.error(request, f'Instrument Tag "{asset_tag}" is already in use.')
            return redirect('instruments:instrument_edit', pk=pk)

        obj.asset_tag = asset_tag
        obj.serial_number = request.POST.get('serial_number', '').strip()
        obj.description = request.POST.get('description', '').strip()
        obj.manufacturer = request.POST.get('manufacturer', '').strip()
        obj.model_number = request.POST.get('model_number', '').strip()
        obj.location = request.POST.get('location', '').strip()
        obj.client_id = request.POST.get('client') or None
        obj.category_id = request.POST.get('category') or None
        obj.calibration_interval_days = int(request.POST.get('calibration_interval_days', 365))
        obj.status = request.POST.get('status', obj.status)
        obj.notes = request.POST.get('notes', '').strip()
        purchase_date = request.POST.get('purchase_date', '').strip()
        obj.purchase_date = purchase_date or None
        obj.save()
        messages.success(request, f'{obj.asset_tag} updated successfully.')
        return redirect('instruments:instrument_detail', pk=pk)

    clients = Client.objects.filter(is_active=True)
    categories = InstrumentCategory.objects.all()
    return render(request, 'instruments/instrument_edit.html', {
        'instrument': obj,
        'clients': clients,
        'categories': categories,
        'status_choices': Instrument.Status.choices,
    })


@login_required
def instrument_create(request):
    if request.method == 'POST':
        category_id = request.POST.get('category') or None
        category = InstrumentCategory.objects.filter(pk=category_id).first() if category_id else None

        # Auto-generate tag from category code; fall back to manual input
        asset_tag = request.POST.get('asset_tag', '').strip()
        if category and category.code:
            asset_tag = category.next_tag()
        elif not asset_tag:
            messages.error(request, 'Instrument Tag is required when no category code is set.')
            clients = Client.objects.filter(is_active=True)
            categories = InstrumentCategory.objects.all()
            return render(request, 'instruments/instrument_create.html', {
                'clients': clients,
                'categories': categories,
            })

        if Instrument.objects.filter(asset_tag=asset_tag).exists():
            # Race condition: regenerate by trying once more
            if category and category.code:
                asset_tag = category.next_tag()
            if Instrument.objects.filter(asset_tag=asset_tag).exists():
                messages.error(request, f'Tag "{asset_tag}" is already taken. Please try again.')
                return redirect('instruments:instrument_create')

        Instrument.objects.create(
            asset_tag=asset_tag,
            serial_number=request.POST['serial_number'],
            description=request.POST['description'],
            manufacturer=request.POST.get('manufacturer', ''),
            model_number=request.POST.get('model_number', ''),
            client_id=request.POST.get('client') or None,
            category=category,
            calibration_interval_days=int(request.POST.get('calibration_interval_days', 365)),
            notes=request.POST.get('notes', ''),
            status=Instrument.Status.ACTIVE,
            created_by=request.user,
        )
        messages.success(request, f'Instrument {asset_tag} registered successfully.')
        return redirect('instruments:instrument_list')

    clients = Client.objects.filter(is_active=True)
    categories = InstrumentCategory.objects.all()
    # Build category-code map for the template JS
    import json
    cat_codes = {str(c.pk): c.code for c in categories if c.code}
    return render(request, 'instruments/instrument_create.html', {
        'clients': clients,
        'categories': categories,
        'cat_codes_json': json.dumps(cat_codes),
    })


@login_required
def instrument_bulk_delete(request):
    """Delete selected instruments that have no calibration jobs."""
    if request.method != 'POST':
        return redirect('instruments:instrument_list')
    ids = request.POST.getlist('selected_ids')
    if not ids:
        messages.warning(request, 'No instruments selected.')
        return redirect('instruments:instrument_list')
    instruments = Instrument.objects.filter(pk__in=ids)
    deleted, skipped = [], []
    for instr in instruments:
        if instr.calibration_jobs.exists():
            skipped.append(instr.asset_tag)
        else:
            deleted.append(instr.asset_tag)
            instr.delete()
    if deleted:
        messages.success(request, f'Deleted: {", ".join(deleted)}.')
    if skipped:
        messages.warning(request, f'Skipped (have calibration history): {", ".join(skipped)}.')
    return redirect('instruments:instrument_list')


@login_required
def instrument_duplicate(request, pk):
    """Create a copy of an instrument with a new asset tag, serial number, and optional different client."""
    original = get_object_or_404(Instrument, pk=pk)
    if request.method == 'POST':
        asset_tag = request.POST.get('asset_tag', '').strip()
        serial_number = request.POST.get('serial_number', '').strip()
        client_id = request.POST.get('client') or None
        if not asset_tag or not serial_number:
            messages.error(request, 'Instrument Tag and Serial Number are required.')
            return redirect('instruments:instrument_list')
        if Instrument.objects.filter(asset_tag=asset_tag).exists():
            messages.error(request, f'Instrument Tag "{asset_tag}" is already in use.')
            return redirect('instruments:instrument_list')
        Instrument.objects.create(
            asset_tag=asset_tag,
            serial_number=serial_number,
            description=original.description,
            manufacturer=original.manufacturer,
            model_number=original.model_number,
            category=original.category,
            client_id=client_id,
            location=original.location,
            calibration_interval_days=original.calibration_interval_days,
            notes=original.notes,
            status=Instrument.Status.ACTIVE,
            created_by=request.user,
        )
        messages.success(request, f'Instrument {asset_tag} created as a copy of {original.asset_tag}.')
        return redirect('instruments:instrument_list')
    return redirect('instruments:instrument_list')


@login_required
def instrument_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    qs, summary, today = _build_export_qs(request)
    sort_label = {
        'tag': 'Instrument Tag', 'client': 'Customer',
        'due_date': 'Next Calibration Date', 'status': 'Status', 'description': 'Description',
    }.get(request.GET.get('sort', 'tag'), 'Instrument Tag')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Instrument List'

    # ── Styles ───────────────────────────────────────────────────
    hdr_fill  = PatternFill('solid', fgColor='1E3A5F')
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    info_font = Font(italic=True, color='555555', size=9)
    red_fill  = PatternFill('solid', fgColor='FEE2E2')
    orange_fill = PatternFill('solid', fgColor='FFF3CD')
    green_fill  = PatternFill('solid', fgColor='DCFCE7')
    thin_border = Border(
        bottom=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
    )

    # ── Info rows ────────────────────────────────────────────────
    ws.append(['CalLIMS — Instrument List Report'])
    ws['A1'].font = Font(bold=True, size=13, color='1E3A5F')
    ws.append([f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}   |   Filter: {", ".join(summary)}   |   Sorted by: {sort_label}'])
    ws['A2'].font = info_font
    ws.append([f'Total records: {qs.count()}'])
    ws['A3'].font = info_font
    ws.append([])  # blank row

    # ── Column headers ────────────────────────────────────────────
    columns = [
        'Instrument Tag', 'Description', 'Serial #', 'Manufacturer', 'Model',
        'Category', 'Customer', 'Status', 'Location',
        'Cal. Interval (days)', 'Last Calibrated', 'Next Due', 'Overdue',
    ]
    ws.append(columns)
    hdr_row = ws.max_row
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(row=hdr_row, column=col_idx)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[hdr_row].height = 20

    # ── Data rows ─────────────────────────────────────────────────
    for instr in qs:
        is_overdue = instr.next_calibration_date and instr.next_calibration_date < today
        due_soon = (
            instr.next_calibration_date
            and not is_overdue
            and instr.next_calibration_date <= today + timezone.timedelta(days=30)
        )
        row = [
            instr.asset_tag,
            instr.description,
            instr.serial_number,
            instr.manufacturer or '',
            instr.model_number or '',
            str(instr.category) if instr.category else '',
            str(instr.client) if instr.client else 'Internal',
            instr.get_status_display(),
            instr.location or '',
            instr.calibration_interval_days,
            instr.last_calibration_date.strftime('%Y-%m-%d') if instr.last_calibration_date else '',
            instr.next_calibration_date.strftime('%Y-%m-%d') if instr.next_calibration_date else '',
            'YES' if is_overdue else '',
        ]
        ws.append(row)
        data_row = ws.max_row
        if is_overdue:
            fill = red_fill
        elif due_soon:
            fill = orange_fill
        elif instr.status == 'CALIBRATED':
            fill = green_fill
        else:
            fill = None
        if fill:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=data_row, column=col_idx).fill = fill

    # ── Column widths ─────────────────────────────────────────────
    col_widths = [16, 30, 16, 18, 16, 16, 20, 18, 18, 12, 14, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    # ── Merge info rows across all columns ────────────────────────
    for row_n in [1, 2, 3]:
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=len(columns))

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'instruments_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def instrument_export_pdf(request):
    from django.template.loader import render_to_string
    from django.http import HttpResponse

    qs, summary, today = _build_export_qs(request)
    instruments = list(qs)
    overdue_count = sum(1 for i in instruments if i.next_calibration_date and i.next_calibration_date < today)
    due_30_count = sum(
        1 for i in instruments
        if i.next_calibration_date
        and not i.next_calibration_date < today
        and i.next_calibration_date <= today + timezone.timedelta(days=30)
    )

    calibrated_count = sum(1 for i in instruments if i.status == 'CALIBRATED')

    html_string = render_to_string('instruments/export_pdf.html', {
        'instruments': instruments,
        'summary': summary,
        'today': today,
        'generated_at': timezone.now(),
        'overdue_count': overdue_count,
        'due_30_count': due_30_count,
        'calibrated_count': calibrated_count,
        'sort_label': {
            'tag': 'Instrument Tag', 'client': 'Customer',
            'due_date': 'Next Calibration Date', 'status': 'Status', 'description': 'Description',
        }.get(request.GET.get('sort', 'tag'), 'Instrument Tag'),
    })

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    except Exception as e:
        return HttpResponse(f'PDF generation failed: {e}', status=500, content_type='text/plain')

    filename = f'instruments_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def sticker_pdf(request, pk):
    import io, base64
    import qrcode
    from django.template.loader import render_to_string
    from django.http import HttpResponse

    obj = get_object_or_404(Instrument.objects.select_related('client', 'category'), pk=pk)

    cert = None
    latest_job = obj.calibration_jobs.filter(status='completed').select_related(
        'method', 'assigned_to'
    ).order_by('-completed_date').first()
    if latest_job:
        try:
            cert = latest_job.certificate
        except Exception:
            cert = None

    size = request.GET.get('size', 'medium')

    if cert:
        qr_url = request.build_absolute_uri(f'/certificates/{cert.pk}/verify/')
    else:
        qr_url = request.build_absolute_uri(f'/instruments/{pk}/')

    qr = qrcode.QRCode(box_size=4, border=2)
    qr.add_data(qr_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    qr_img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    html_string = render_to_string('instruments/sticker_pdf.html', {
        'instrument': obj,
        'cert': cert,
        'job': latest_job,
        'qr_b64': qr_b64,
        'size': size,
        'lab_name': 'CalLIMS Calibration Laboratory',
    })

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    except Exception as e:
        return HttpResponse(f'PDF generation failed: {e}', status=500, content_type='text/plain')

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="sticker-{obj.asset_tag}.pdf"'
    return response
